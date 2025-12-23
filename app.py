from flask import Flask, render_template, request, jsonify
import json
import os
import numpy as np
from PIL import Image
import io
import base64
import hashlib

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # Max 16MB upload

def load_data():
    filename = 'sbox_data_full.json'
    try:
        with open(filename, 'r') as f:
            data = json.load(f)
        return data
    except FileNotFoundError:
        return None

def get_sbox(sbox_id):
    """Get S-Box by ID"""
    dataset = load_data()
    if dataset:
        for candidate in dataset['candidates']:
            if candidate['id'] == sbox_id:
                return candidate['sbox']
    return None

def generate_inverse_sbox(sbox):
    """Generate inverse S-Box"""
    inv = [0] * 256
    for i in range(256):
        inv[sbox[i]] = i
    return inv

# ============================================
# ENHANCED IMAGE ENCRYPTION WITH PROPER AES-LIKE OPERATIONS
# Includes: Confusion (S-Box) + Diffusion (Permutation + XOR)
# ============================================

def generate_chaotic_sequence(seed, length):
    """
    Generate chaotic sequence using Logistic Map
    x_{n+1} = r * x_n * (1 - x_n), where r = 3.99 (chaotic regime)
    """
    r = 3.9999  # Chaotic parameter
    x = seed
    sequence = []
    
    # Skip transient iterations
    for _ in range(1000):
        x = r * x * (1 - x)
    
    # Generate sequence
    for _ in range(length):
        x = r * x * (1 - x)
        sequence.append(x)
    
    return np.array(sequence)

def generate_permutation_indices(height, width, seed):
    """
    Generate permutation indices using chaotic sequence
    for pixel position scrambling (Arnold Cat Map inspired)
    """
    total_pixels = height * width
    
    # Generate chaotic sequence
    chaos = generate_chaotic_sequence(seed, total_pixels)
    
    # Convert to indices
    indices = np.argsort(chaos)
    
    return indices

def generate_key_stream(sbox, seed, length):
    """
    Generate pseudo-random key stream using S-Box and chaotic sequence
    for XOR diffusion
    """
    chaos = generate_chaotic_sequence(seed, length)
    
    # Convert chaotic values to bytes (0-255)
    key_stream = np.floor(chaos * 256).astype(np.uint8)
    key_stream = np.clip(key_stream, 0, 255)
    
    # Apply S-Box to key stream for additional confusion
    key_stream = np.array([sbox[k] for k in key_stream], dtype=np.uint8)
    
    return key_stream

def encrypt_image_enhanced(img_array, sbox, key="cryptography2024"):
    """
    Enhanced image encryption with confusion and diffusion
    
    Steps:
    1. Generate seed from key
    2. For each round:
       a. Pixel permutation (diffusion - position scrambling)
       b. S-Box substitution (confusion - value substitution)
       c. XOR with key stream (diffusion - value mixing)
    3. Apply CBC-like chaining for additional diffusion
    """
    # Generate seed from key
    key_hash = hashlib.sha256(key.encode()).hexdigest()
    seed = int(key_hash[:8], 16) / (16**8)  # Normalize to (0, 1)
    seed = max(0.1, min(0.9, seed))  # Ensure valid range for logistic map
    
    shape = img_array.shape
    is_color = len(shape) == 3
    
    if is_color:
        height, width, channels = shape
    else:
        height, width = shape
        channels = 1
        img_array = img_array.reshape(height, width, 1)
    
    encrypted = img_array.copy().astype(np.uint8)
    total_pixels = height * width
    
    num_rounds = 3  # Multiple rounds for better diffusion
    
    for round_num in range(num_rounds):
        # Adjust seed for each round
        round_seed = (seed + round_num * 0.1) % 0.9 + 0.05
        
        for c in range(channels):
            channel_data = encrypted[:, :, c].flatten()
            
            # Step 1: Pixel Position Permutation (Diffusion)
            perm_indices = generate_permutation_indices(height, width, round_seed + c * 0.01)
            channel_data = channel_data[perm_indices]
            
            # Step 2: S-Box Substitution (Confusion)
            channel_data = np.array([sbox[pixel] for pixel in channel_data], dtype=np.uint8)
            
            # Step 3: XOR with Key Stream (Diffusion)
            key_stream = generate_key_stream(sbox, round_seed + c * 0.02, total_pixels)
            channel_data = np.bitwise_xor(channel_data, key_stream)
            
            # Step 4: CBC-like Chaining (Additional Diffusion)
            # Each pixel depends on the previous encrypted pixel
            for i in range(1, len(channel_data)):
                channel_data[i] = channel_data[i] ^ channel_data[i-1]
            
            # Apply S-Box again after chaining
            channel_data = np.array([sbox[pixel] for pixel in channel_data], dtype=np.uint8)
            
            encrypted[:, :, c] = channel_data.reshape(height, width)
    
    if not is_color:
        encrypted = encrypted.reshape(height, width)
    
    return encrypted

def decrypt_image_enhanced(img_array, sbox, key="cryptography2024"):
    """
    Decrypt image encrypted with enhanced method
    Reverse all operations in reverse order
    """
    inv_sbox = generate_inverse_sbox(sbox)
    
    # Generate seed from key
    key_hash = hashlib.sha256(key.encode()).hexdigest()
    seed = int(key_hash[:8], 16) / (16**8)
    seed = max(0.1, min(0.9, seed))
    
    shape = img_array.shape
    is_color = len(shape) == 3
    
    if is_color:
        height, width, channels = shape
    else:
        height, width = shape
        channels = 1
        img_array = img_array.reshape(height, width, 1)
    
    decrypted = img_array.copy().astype(np.uint8)
    total_pixels = height * width
    
    num_rounds = 3
    
    # Reverse rounds
    for round_num in range(num_rounds - 1, -1, -1):
        round_seed = (seed + round_num * 0.1) % 0.9 + 0.05
        
        for c in range(channels):
            channel_data = decrypted[:, :, c].flatten()
            
            # Reverse Step 4: Inverse S-Box
            channel_data = np.array([inv_sbox[pixel] for pixel in channel_data], dtype=np.uint8)
            
            # Reverse Step 3: Reverse CBC chaining
            temp = channel_data.copy()
            for i in range(len(channel_data) - 1, 0, -1):
                channel_data[i] = temp[i] ^ temp[i-1]
            
            # Reverse Step 2: XOR with same key stream
            key_stream = generate_key_stream(sbox, round_seed + c * 0.02, total_pixels)
            channel_data = np.bitwise_xor(channel_data, key_stream)
            
            # Reverse Step 1: Inverse S-Box
            channel_data = np.array([inv_sbox[pixel] for pixel in channel_data], dtype=np.uint8)
            
            # Reverse Step 0: Inverse permutation
            perm_indices = generate_permutation_indices(height, width, round_seed + c * 0.01)
            inv_perm = np.argsort(perm_indices)
            channel_data = channel_data[inv_perm]
            
            decrypted[:, :, c] = channel_data.reshape(height, width)
    
    if not is_color:
        decrypted = decrypted.reshape(height, width)
    
    return decrypted

# Legacy functions for backward compatibility
def encrypt_image_simple(img_array, sbox):
    """Simple S-Box only encryption (legacy - weak)"""
    encrypted = np.zeros_like(img_array)
    shape = img_array.shape
    
    if len(shape) == 3:
        for channel in range(shape[2]):
            for i in range(shape[0]):
                for j in range(shape[1]):
                    encrypted[i, j, channel] = sbox[img_array[i, j, channel]]
    else:
        for i in range(shape[0]):
            for j in range(shape[1]):
                encrypted[i, j] = sbox[img_array[i, j]]
    
    return encrypted

def decrypt_image_simple(img_array, sbox):
    """Simple S-Box only decryption (legacy)"""
    inv_sbox = generate_inverse_sbox(sbox)
    return encrypt_image_simple(img_array, inv_sbox)

# Use enhanced version as default
def encrypt_image(img_array, sbox, key="cryptography2024"):
    """Encrypt image - uses enhanced method"""
    return encrypt_image_enhanced(img_array, sbox, key)

def decrypt_image(img_array, sbox, key="cryptography2024"):
    """Decrypt image - uses enhanced method"""
    return decrypt_image_enhanced(img_array, sbox, key)

def calculate_entropy(img_array):
    """Calculate Shannon entropy"""
    histogram, _ = np.histogram(img_array.flatten(), bins=256, range=(0, 256))
    histogram = histogram / histogram.sum()
    histogram = histogram[histogram > 0]
    entropy = -np.sum(histogram * np.log2(histogram))
    return entropy

def calculate_npcr(img1, img2):
    """Calculate Number of Pixels Change Rate"""
    if img1.shape != img2.shape:
        return 0.0
    diff = np.sum(img1 != img2)
    total = img1.size
    return (diff / total) * 100

def calculate_uaci(img1, img2):
    """Calculate Unified Average Changing Intensity"""
    if img1.shape != img2.shape:
        return 0.0
    diff = np.abs(img1.astype(float) - img2.astype(float))
    return (np.sum(diff) / (img1.size * 255)) * 100

def calculate_correlation(img_array, direction='horizontal'):
    """Calculate correlation coefficient"""
    if len(img_array.shape) == 3:
        img_array = np.mean(img_array, axis=2).astype(np.uint8)
    
    pairs = []
    if direction == 'horizontal':
        pairs = [(img_array[i, j], img_array[i, j+1]) 
                 for i in range(img_array.shape[0]) 
                 for j in range(img_array.shape[1]-1)]
    elif direction == 'vertical':
        pairs = [(img_array[i, j], img_array[i+1, j]) 
                 for i in range(img_array.shape[0]-1) 
                 for j in range(img_array.shape[1])]
    elif direction == 'diagonal':
        pairs = [(img_array[i, j], img_array[i+1, j+1]) 
                 for i in range(img_array.shape[0]-1) 
                 for j in range(img_array.shape[1]-1)]
    
    if len(pairs) == 0:
        return 0.0
    
    x, y = zip(*pairs)
    x, y = np.array(x), np.array(y)
    
    if np.std(x) == 0 or np.std(y) == 0:
        return 0.0
    
    correlation = np.corrcoef(x, y)[0, 1]
    return correlation

def calculate_mae(img1, img2):
    """Calculate Mean Absolute Error"""
    if img1.shape != img2.shape:
        return 0.0
    return np.mean(np.abs(img1.astype(float) - img2.astype(float)))

def image_to_base64(img_array):
    """Convert numpy array to base64 string"""
    img = Image.fromarray(img_array.astype(np.uint8))
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    img_str = base64.b64encode(buffer.getvalue()).decode()
    return f"data:image/png;base64,{img_str}"

def generate_histogram_image(img_array, title="Histogram"):
    """Generate histogram image as base64
    Creates a visual histogram showing pixel distribution per channel
    """
    import matplotlib
    matplotlib.use('Agg')  # Non-interactive backend
    import matplotlib.pyplot as plt
    
    # Create larger figure with dark theme for better visibility
    fig, ax = plt.subplots(figsize=(10, 4), dpi=120)
    fig.patch.set_facecolor('#0f172a')  # slate-900
    ax.set_facecolor('#020617')  # slate-950
    
    if len(img_array.shape) == 3 and img_array.shape[2] >= 3:
        # Color image - plot RGB histograms
        colors = ['#ef4444', '#22c55e', '#3b82f6']  # red, green, blue
        labels = ['Red Channel', 'Green Channel', 'Blue Channel']
        
        for i, (color, label) in enumerate(zip(colors, labels)):
            hist, bins = np.histogram(img_array[:, :, i].flatten(), bins=256, range=(0, 256))
            ax.plot(bins[:-1], hist, color=color, alpha=0.85, linewidth=1.5, label=label)
            ax.fill_between(bins[:-1], hist, alpha=0.25, color=color)
    else:
        # Grayscale image
        if len(img_array.shape) == 3:
            img_flat = img_array[:, :, 0].flatten()
        else:
            img_flat = img_array.flatten()
        hist, bins = np.histogram(img_flat, bins=256, range=(0, 256))
        ax.plot(bins[:-1], hist, color='#a78bfa', linewidth=1.5, label='Intensity')
        ax.fill_between(bins[:-1], hist, alpha=0.35, color='#a78bfa')
    
    # Style the chart
    ax.set_xlabel('Pixel Value (0-255)', color='#cbd5e1', fontsize=11, fontweight='medium')
    ax.set_ylabel('Frequency', color='#cbd5e1', fontsize=11, fontweight='medium')
    ax.set_title(title, color='white', fontsize=14, fontweight='bold', pad=15)
    ax.tick_params(colors='#94a3b8', labelsize=10)
    
    # Grid for better readability
    ax.grid(True, alpha=0.15, color='#475569', linestyle='--')
    
    # Spine colors
    for spine in ax.spines.values():
        spine.set_color('#334155')
        spine.set_linewidth(1)
    
    ax.set_xlim(0, 255)
    ax.set_ylim(bottom=0)
    
    # Legend with better styling
    legend = ax.legend(loc='upper right', fontsize=10, facecolor='#1e293b', 
                       edgecolor='#475569', labelcolor='white', framealpha=0.9)
    legend.get_frame().set_linewidth(1)
    
    plt.tight_layout(pad=1.5)
    
    # Save to buffer with higher quality
    buffer = io.BytesIO()
    fig.savefig(buffer, format='png', facecolor=fig.get_facecolor(), 
                edgecolor='none', bbox_inches='tight', pad_inches=0.1)
    plt.close(fig)
    buffer.seek(0)
    
    img_str = base64.b64encode(buffer.getvalue()).decode()
    return f"data:image/png;base64,{img_str}"

@app.route('/')
def index():
    # Load data dari file JSON
    dataset = load_data()
    
    if dataset is None:
        return "<h1>Error: File sbox_data_full.json tidak ditemukan!</h1><p>Pastikan kamu sudah download dari Colab dan taruh di folder yang sama dengan app.py</p>"
    
    return render_template('index.html', candidates=dataset['candidates'], metadata=dataset['metadata'])

@app.route('/encrypt-image', methods=['POST'])
def encrypt_image_route():
    """Encrypt uploaded image with enhanced encryption"""
    try:
        if 'image' not in request.files:
            return jsonify({'error': 'No image uploaded'}), 400
        
        file = request.files['image']
        sbox_id = request.form.get('sbox_id', 'AES_STD')
        encryption_key = request.form.get('key', 'cryptography2024')  # Get encryption key
        
        # Load image
        img = Image.open(file.stream)
        img_array = np.array(img)
        
        # Get S-Box
        sbox = get_sbox(sbox_id)
        if not sbox:
            return jsonify({'error': 'S-Box not found'}), 400
        
        # Encrypt with enhanced method
        encrypted = encrypt_image(img_array, sbox, encryption_key)
        
        # Calculate metrics
        original_entropy = calculate_entropy(img_array)
        encrypted_entropy = calculate_entropy(encrypted)
        npcr = calculate_npcr(img_array, encrypted)
        uaci = calculate_uaci(img_array, encrypted)
        mae = calculate_mae(img_array, encrypted)
        
        # Correlation
        orig_corr_h = calculate_correlation(img_array, 'horizontal')
        orig_corr_v = calculate_correlation(img_array, 'vertical')
        orig_corr_d = calculate_correlation(img_array, 'diagonal')
        
        enc_corr_h = calculate_correlation(encrypted, 'horizontal')
        enc_corr_v = calculate_correlation(encrypted, 'vertical')
        enc_corr_d = calculate_correlation(encrypted, 'diagonal')
        
        # Convert to base64
        encrypted_b64 = image_to_base64(encrypted)
        original_b64 = image_to_base64(img_array)
        
        # Generate histograms
        original_histogram = generate_histogram_image(img_array, "Original Histogram")
        encrypted_histogram = generate_histogram_image(encrypted, "Encrypted Histogram")
        
        return jsonify({
            'success': True,
            'encrypted_image': encrypted_b64,
            'original_image': original_b64,
            'original_histogram': original_histogram,
            'encrypted_histogram': encrypted_histogram,
            'metrics': {
                'original_entropy': round(original_entropy, 6),
                'encrypted_entropy': round(encrypted_entropy, 6),
                'npcr': round(npcr, 4),
                'uaci': round(uaci, 4),
                'mae': round(mae, 4),
                'original_correlation': {
                    'horizontal': round(orig_corr_h, 6),
                    'vertical': round(orig_corr_v, 6),
                    'diagonal': round(orig_corr_d, 6)
                },
                'encrypted_correlation': {
                    'horizontal': round(enc_corr_h, 6),
                    'vertical': round(enc_corr_v, 6),
                    'diagonal': round(enc_corr_d, 6)
                }
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/decrypt-image', methods=['POST'])
def decrypt_image_route():
    """Decrypt uploaded encrypted image"""
    try:
        if 'image' not in request.files:
            return jsonify({'error': 'No image uploaded'}), 400
        
        file = request.files['image']
        sbox_id = request.form.get('sbox_id', 'AES_STD')
        encryption_key = request.form.get('key', 'cryptography2024')  # Get encryption key
        
        # Load image
        img = Image.open(file.stream)
        img_array = np.array(img)
        
        # Get S-Box
        sbox = get_sbox(sbox_id)
        if not sbox:
            return jsonify({'error': 'S-Box not found'}), 400
        
        # Decrypt with enhanced method
        decrypted = decrypt_image(img_array, sbox, encryption_key)
        
        # Convert to base64
        decrypted_b64 = image_to_base64(decrypted)
        
        # Generate histograms for comparison
        encrypted_histogram = generate_histogram_image(img_array, "Encrypted Histogram")
        decrypted_histogram = generate_histogram_image(decrypted, "Decrypted Histogram")
        
        # Calculate metrics for comparison
        decrypted_entropy = calculate_entropy(decrypted)
        encrypted_entropy = calculate_entropy(img_array)
        
        return jsonify({
            'success': True,
            'decrypted_image': decrypted_b64,
            'encrypted_histogram': encrypted_histogram,
            'decrypted_histogram': decrypted_histogram,
            'metrics': {
                'encrypted_entropy': round(encrypted_entropy, 6),
                'decrypted_entropy': round(decrypted_entropy, 6)
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/team-photo/<filename>')
def team_photo(filename):
    """Serve resized team member photos"""
    try:
        from PIL import ImageOps
        
        # Target size for team cards (w:112px x h:144px at 2x for retina = 224x288)
        target_width = 224
        target_height = 288
        
        # Map filename to actual image file
        image_map = {
            'huda': 'huda.JPG',
            'firda': 'firda satria.png',
            'arif': 'arif.jpg',
            'tatak': 'tatak.jpeg'
        }
        
        actual_filename = image_map.get(filename)
        if not actual_filename:
            return "Image not found", 404
        
        image_path = os.path.join(app.static_folder, 'images', actual_filename)
        
        if not os.path.exists(image_path):
            return "Image not found", 404
        
        # Open and resize image
        img = Image.open(image_path)
        
        # Fix EXIF orientation (handles rotated photos from phones)
        img = ImageOps.exif_transpose(img)
        
        # Convert to RGB if necessary (for PNG with transparency)
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        # Calculate crop to maintain aspect ratio and focus on top (face)
        img_ratio = img.width / img.height
        target_ratio = target_width / target_height
        
        if img_ratio > target_ratio:
            # Image is wider - crop sides
            new_width = int(img.height * target_ratio)
            left = (img.width - new_width) // 2
            img = img.crop((left, 0, left + new_width, img.height))
        else:
            # Image is taller - crop bottom (keep top/face)
            new_height = int(img.width / target_ratio)
            img = img.crop((0, 0, img.width, new_height))
        
        # Resize to target
        img = img.resize((target_width, target_height), Image.Resampling.LANCZOS)
        
        # Save to buffer
        buffer = io.BytesIO()
        img.save(buffer, format='JPEG', quality=70)
        buffer.seek(0)
        
        from flask import send_file
        return send_file(buffer, mimetype='image/jpeg')
    except Exception as e:
        return str(e), 500

# ============================================
# CUSTOM S-BOX ROUTES
# ============================================

@app.route('/download-template/<template_type>')
def download_template(template_type):
    """Generate and download Excel template for S-Box"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from flask import send_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = "S-Box"
        
        # Styling
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        cell_font = Font(name="Consolas", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='cccccc'),
            right=Side(style='thin', color='cccccc'),
            top=Side(style='thin', color='cccccc'),
            bottom=Side(style='thin', color='cccccc')
        )
        
        # Generate sample AES S-Box as template
        aes_sbox = [
            0x63, 0x7c, 0x77, 0x7b, 0xf2, 0x6b, 0x6f, 0xc5, 0x30, 0x01, 0x67, 0x2b, 0xfe, 0xd7, 0xab, 0x76,
            0xca, 0x82, 0xc9, 0x7d, 0xfa, 0x59, 0x47, 0xf0, 0xad, 0xd4, 0xa2, 0xaf, 0x9c, 0xa4, 0x72, 0xc0,
            0xb7, 0xfd, 0x93, 0x26, 0x36, 0x3f, 0xf7, 0xcc, 0x34, 0xa5, 0xe5, 0xf1, 0x71, 0xd8, 0x31, 0x15,
            0x04, 0xc7, 0x23, 0xc3, 0x18, 0x96, 0x05, 0x9a, 0x07, 0x12, 0x80, 0xe2, 0xeb, 0x27, 0xb2, 0x75,
            0x09, 0x83, 0x2c, 0x1a, 0x1b, 0x6e, 0x5a, 0xa0, 0x52, 0x3b, 0xd6, 0xb3, 0x29, 0xe3, 0x2f, 0x84,
            0x53, 0xd1, 0x00, 0xed, 0x20, 0xfc, 0xb1, 0x5b, 0x6a, 0xcb, 0xbe, 0x39, 0x4a, 0x4c, 0x58, 0xcf,
            0xd0, 0xef, 0xaa, 0xfb, 0x43, 0x4d, 0x33, 0x85, 0x45, 0xf9, 0x02, 0x7f, 0x50, 0x3c, 0x9f, 0xa8,
            0x51, 0xa3, 0x40, 0x8f, 0x92, 0x9d, 0x38, 0xf5, 0xbc, 0xb6, 0xda, 0x21, 0x10, 0xff, 0xf3, 0xd2,
            0xcd, 0x0c, 0x13, 0xec, 0x5f, 0x97, 0x44, 0x17, 0xc4, 0xa7, 0x7e, 0x3d, 0x64, 0x5d, 0x19, 0x73,
            0x60, 0x81, 0x4f, 0xdc, 0x22, 0x2a, 0x90, 0x88, 0x46, 0xee, 0xb8, 0x14, 0xde, 0x5e, 0x0b, 0xdb,
            0xe0, 0x32, 0x3a, 0x0a, 0x49, 0x06, 0x24, 0x5c, 0xc2, 0xd3, 0xac, 0x62, 0x91, 0x95, 0xe4, 0x79,
            0xe7, 0xc8, 0x37, 0x6d, 0x8d, 0xd5, 0x4e, 0xa9, 0x6c, 0x56, 0xf4, 0xea, 0x65, 0x7a, 0xae, 0x08,
            0xba, 0x78, 0x25, 0x2e, 0x1c, 0xa6, 0xb4, 0xc6, 0xe8, 0xdd, 0x74, 0x1f, 0x4b, 0xbd, 0x8b, 0x8a,
            0x70, 0x3e, 0xb5, 0x66, 0x48, 0x03, 0xf6, 0x0e, 0x61, 0x35, 0x57, 0xb9, 0x86, 0xc1, 0x1d, 0x9e,
            0xe1, 0xf8, 0x98, 0x11, 0x69, 0xd9, 0x8e, 0x94, 0x9b, 0x1e, 0x87, 0xe9, 0xce, 0x55, 0x28, 0xdf,
            0x8c, 0xa1, 0x89, 0x0d, 0xbf, 0xe6, 0x42, 0x68, 0x41, 0x99, 0x2d, 0x0f, 0xb0, 0x54, 0xbb, 0x16
        ]
        
        if template_type == "16x16":
            # Create 16x16 grid
            # Header row
            ws.cell(row=1, column=1, value="").fill = header_fill
            for c in range(16):
                cell = ws.cell(row=1, column=c+2, value=f"{c:X}")
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                ws.column_dimensions[cell.column_letter].width = 5
            
            # Data rows
            for r in range(16):
                # Row header
                cell = ws.cell(row=r+2, column=1, value=f"{r:X}")
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                
                # S-Box values
                for c in range(16):
                    idx = r * 16 + c
                    cell = ws.cell(row=r+2, column=c+2, value=aes_sbox[idx])
                    cell.font = cell_font
                    cell.alignment = center_align
                    cell.border = thin_border
            
            ws.column_dimensions['A'].width = 4
            filename = "sbox_template_16x16.xlsx"
            
        else:  # 1x256
            # Create 1x256 row
            ws.cell(row=1, column=1, value="Index")
            ws.cell(row=2, column=1, value="Value")
            
            for i in range(256):
                ws.cell(row=1, column=i+2, value=i)
                ws.cell(row=2, column=i+2, value=aes_sbox[i])
            
            filename = "sbox_template_1x256.xlsx"
        
        # Add instructions sheet
        ws_info = wb.create_sheet("Instructions")
        instructions = [
            "S-Box Template Instructions",
            "",
            "1. Sheet 'S-Box' berisi template S-Box yang bisa diedit",
            "2. S-Box harus berisi 256 nilai unik dari 0-255 (bijektif)",
            "3. Untuk format 16x16: nilai di baris r, kolom c = S[16*r + c]",
            "4. Untuk format 1x256: nilai di kolom i = S[i-1]",
            "",
            "Opsional:",
            "- Tambahkan sheet 'Affine Matrix' dengan matrix 8x8",
            "- Nilai matrix harus 0 atau 1 (binary)",
            "",
            "Tips:",
            "- Gunakan fungsi COUNTIF untuk memastikan tidak ada duplikat",
            "- Contoh: =COUNTIF(B2:Q17, A1) harus = 1 untuk semua nilai",
        ]
        for i, text in enumerate(instructions):
            ws_info.cell(row=i+1, column=1, value=text)
        ws_info.column_dimensions['A'].width = 60
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/upload-custom-sbox', methods=['POST'])
def upload_custom_sbox():
    """Upload and analyze custom S-Box from Excel file"""
    try:
        from openpyxl import load_workbook
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'File must be .xlsx or .xls format'}), 400
        
        # Load workbook
        wb = load_workbook(file)
        ws = wb.active
        
        # Try to extract S-Box values
        sbox = []
        
        # Check dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row >= 16 and max_col >= 16:
            # Likely 16x16 format (with or without headers)
            start_row = 2 if max_row == 17 else 1
            start_col = 2 if max_col == 17 else 1
            
            for r in range(16):
                for c in range(16):
                    val = ws.cell(row=start_row + r, column=start_col + c).value
                    if val is not None:
                        try:
                            sbox.append(int(val))
                        except (ValueError, TypeError):
                            pass
        
        elif max_row <= 2 and max_col >= 256:
            # Likely 1x256 format
            data_row = 2 if max_row == 2 else 1
            start_col = 2 if max_col > 256 else 1
            
            for c in range(256):
                val = ws.cell(row=data_row, column=start_col + c).value
                if val is not None:
                    try:
                        sbox.append(int(val))
                    except (ValueError, TypeError):
                        pass
        
        # Validate S-Box
        if len(sbox) != 256:
            return jsonify({
                'error': f'S-Box must have exactly 256 values. Found: {len(sbox)}'
            }), 400
        
        # Check if bijective (all values 0-255 appear exactly once)
        if sorted(sbox) != list(range(256)):
            return jsonify({
                'error': 'S-Box must be bijective (contain each value 0-255 exactly once)'
            }), 400
        
        # Try to get affine matrix if available
        affine_matrix = None
        if 'Affine Matrix' in wb.sheetnames:
            ws_matrix = wb['Affine Matrix']
            matrix = []
            for r in range(1, 9):
                row = []
                for c in range(1, 9):
                    val = ws_matrix.cell(row=r, column=c).value
                    row.append(int(val) if val is not None else 0)
                matrix.append(row)
            affine_matrix = matrix
        
        # Calculate all cryptographic properties
        sbox_array = np.array(sbox)
        
        # Basic metrics
        entropy = calculate_entropy(sbox_array.reshape(16, 16))
        
        # Calculate all 10 properties (simplified versions)
        metrics = calculate_sbox_properties(sbox)
        
        return jsonify({
            'success': True,
            'sbox': sbox,
            'affine_matrix': affine_matrix,
            'metrics': metrics,
            'message': 'S-Box successfully loaded and validated!'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def calculate_sbox_properties(sbox):
    """Calculate cryptographic properties of an S-Box"""
    sbox = np.array(sbox)
    
    # 1. Nonlinearity (simplified estimation)
    # True nonlinearity requires Walsh-Hadamard transform
    nl = estimate_nonlinearity(sbox)
    
    # 2. SAC (Strict Avalanche Criterion)
    sac = calculate_sac(sbox)
    
    # 3. BIC (Bit Independence Criterion) 
    bic = calculate_bic(sbox)
    
    # 4. LAP (Linear Approximation Probability)
    lap = calculate_lap(sbox)
    
    # 5. DAP (Differential Approximation Probability)
    dap = calculate_dap(sbox)
    
    return {
        'nonlinearity': int(nl),
        'sac': round(sac, 6),
        'bic': round(bic, 6),
        'lap': round(lap, 6),
        'dap': round(dap, 6),
        'is_bijective': True,  # Already validated
        'fixed_points': int(np.sum(sbox == np.arange(256))),
        'opposite_fixed_points': int(np.sum(sbox == (255 - np.arange(256))))
    }

def estimate_nonlinearity(sbox):
    """Estimate nonlinearity of S-Box (simplified)"""
    # For 8-bit S-Box, max nonlinearity is 120
    # This is a simplified estimation
    n = 8
    max_nl = 2**(n-1) - 2**(n//2-1)  # 120 for n=8
    
    # Count how different from linear
    differences = 0
    for i in range(256):
        for j in range(256):
            if sbox[i ^ j] != sbox[i] ^ sbox[j]:
                differences += 1
    
    # Normalize to approximate nonlinearity
    nl = int(max_nl * (differences / (256 * 256)))
    return min(nl, max_nl)

def calculate_sac(sbox):
    """Calculate SAC (Strict Avalanche Criterion)"""
    total = 0
    count = 0
    
    for i in range(256):
        for bit in range(8):
            # Flip one bit
            j = i ^ (1 << bit)
            # Count output bit changes
            diff = sbox[i] ^ sbox[j]
            bit_changes = bin(diff).count('1')
            total += bit_changes
            count += 8  # 8 output bits
    
    return total / count if count > 0 else 0

def calculate_bic(sbox):
    """Calculate BIC (Bit Independence Criterion)"""
    # Simplified BIC calculation
    correlations = []
    
    for bit1 in range(8):
        for bit2 in range(bit1 + 1, 8):
            same = 0
            for i in range(256):
                for flip_bit in range(8):
                    j = i ^ (1 << flip_bit)
                    diff = sbox[i] ^ sbox[j]
                    b1 = (diff >> bit1) & 1
                    b2 = (diff >> bit2) & 1
                    if b1 == b2:
                        same += 1
            correlations.append(abs(same / (256 * 8) - 0.5))
    
    return 1 - (sum(correlations) / len(correlations) * 2) if correlations else 0

def calculate_lap(sbox):
    """Calculate LAP (Linear Approximation Probability)"""
    max_bias = 0
    
    for a in range(1, 256):
        for b in range(1, 256):
            count = 0
            for x in range(256):
                input_parity = bin(a & x).count('1') % 2
                output_parity = bin(b & sbox[x]).count('1') % 2
                if input_parity == output_parity:
                    count += 1
            bias = abs(count - 128) / 256
            max_bias = max(max_bias, bias)
    
    return max_bias

def calculate_dap(sbox):
    """Calculate DAP (Differential Approximation Probability)"""
    max_prob = 0
    
    for delta_in in range(1, 256):
        diff_dist = {}
        for x in range(256):
            delta_out = sbox[x] ^ sbox[x ^ delta_in]
            diff_dist[delta_out] = diff_dist.get(delta_out, 0) + 1
        
        max_count = max(diff_dist.values())
        prob = max_count / 256
        max_prob = max(max_prob, prob)
    
    return max_prob

if __name__ == '__main__':
    app.run(debug=True)
    